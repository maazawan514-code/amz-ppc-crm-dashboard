"""
master_kw_parser.py
---------------------
Parses the "Master Keyword Research" template: a keyword x competitor-ASIN
organic rank grid (Keyword Phrase | Search Volume | Relevancy | ASIN1 | ASIN2 | ...),
with a Main Keyword / Roots block above it. This is the format used for the
niche-entry decision method (green/orange/red competitor classification).

We read it directly with openpyxl (not pandas) because the layout is
positional (metadata rows above a shifting header row), not a clean table.
"""

import openpyxl
import pandas as pd
import numpy as np


import re


ASIN_PATTERN = re.compile(r"^B[0-9A-Z]{9}$")


def parse_raw_cerebro_asin_export(filepath_or_buffer):
    """
    Parses a raw Helium10 Cerebro export that was generated with 'Add ASINs to
    track' turned on — the standard Cerebro columns plus one column per
    tracked competitor ASIN (organic rank, '-' if not ranking). This is a
    single-source automated equivalent of the manual Master Keyword Research
    grid — no manual compiling needed.

    Returns the same shape as parse_master_kw_file: dict(main_keyword, roots, grid)
    """
    if hasattr(filepath_or_buffer, "name") and str(filepath_or_buffer.name).lower().endswith(".csv"):
        df = pd.read_csv(filepath_or_buffer)
    elif isinstance(filepath_or_buffer, str) and filepath_or_buffer.lower().endswith(".csv"):
        df = pd.read_csv(filepath_or_buffer)
    else:
        try:
            df = pd.read_csv(filepath_or_buffer)
        except Exception:
            filepath_or_buffer.seek(0) if hasattr(filepath_or_buffer, "seek") else None
            df = pd.read_excel(filepath_or_buffer)

    df.columns = [str(c).strip() for c in df.columns]

    kw_col = next((c for c in df.columns if c.strip().lower() == "keyword phrase"), None)
    sv_col = next((c for c in df.columns if c.strip().lower() == "search volume"), None)
    iq_col = next((c for c in df.columns if "cerebro iq" in c.strip().lower()), None)
    rel_col = next((c for c in df.columns if c.strip().lower() == "relevancy"), None)

    if kw_col is None:
        raise ValueError(f"Could not find 'Keyword Phrase' column. Detected headers: {list(df.columns)}")

    asin_cols = [c for c in df.columns if ASIN_PATTERN.match(c.strip())]
    if not asin_cols:
        raise ValueError(
            "Is file mein koi tracked ASIN column nahi mila (Helium10 'Add ASINs to track' feature "
            "se export karein — tab har competitor ASIN ka apna column hoga)."
        )

    keep = {kw_col: "keyword_phrase", sv_col: "search_volume"}
    base = df.rename(columns=keep)[["keyword_phrase", "search_volume"]].copy()
    base["search_volume"] = base["search_volume"].map(_clean_num)

    asin_rank_matrix = df[asin_cols].map(_clean_num)

    if rel_col:
        base["relevancy"] = df[rel_col].map(_clean_num)
    else:
        # Auto-calculate: COUNTIF-equivalent — how many tracked competitor
        # ASINs rank < 30 for this keyword (matches the manual formula:
        # =COUNTIF(competitor_rank_columns, "<30"))
        base["relevancy"] = (asin_rank_matrix < 30).sum(axis=1).astype(float)

    long_rows = []
    for asin in asin_cols:
        ranks = df[asin].map(_clean_num)
        chunk = pd.DataFrame({
            "keyword_phrase": base["keyword_phrase"],
            "search_volume": base["search_volume"],
            "relevancy": base["relevancy"],
            "asin": asin,
            "organic_rank": ranks,
        })
        long_rows.append(chunk)

    grid = pd.concat(long_rows, ignore_index=True)
    grid = grid[grid["keyword_phrase"].notna() & (grid["keyword_phrase"].astype(str).str.strip() != "")]

    return {"main_keyword": None, "roots": [], "grid": grid}


def _clean_num(v):
    if v is None:
        return np.nan
    if isinstance(v, (int, float)):
        return float(v)
    s = str(v).strip()
    if s in ("-", "", "N/A", "n/a"):
        return np.nan
    s = s.replace(",", "")
    try:
        return float(s)
    except ValueError:
        return np.nan


def parse_master_kw_file(filepath_or_buffer):
    """
    Returns: dict(main_keyword, roots, grid) where grid is a long-format
    DataFrame: keyword_phrase, search_volume, relevancy, asin, organic_rank
    """
    wb = openpyxl.load_workbook(filepath_or_buffer, data_only=True)
    ws = wb[wb.sheetnames[0]]

    header_row = None
    main_keyword = None
    roots = []

    for r in range(1, ws.max_row + 1):
        a_val = ws.cell(row=r, column=1).value
        if a_val is None:
            continue
        a_str = str(a_val).strip().lower()
        if a_str == "keyword phrase":
            header_row = r
            break
        if a_str == "main keyword":
            main_keyword = ws.cell(row=r, column=2).value
        if a_str == "roots":
            c = 2
            while ws.cell(row=r, column=c).value not in (None, ""):
                roots.append(ws.cell(row=r, column=c).value)
                c += 1

    if header_row is None:
        raise ValueError("Could not find a 'Keyword Phrase' header row — is this the right file format?")

    # Column 1 = Keyword Phrase, 2 = Search Volume, 3 = Relevancy, 4+ = ASIN columns
    max_col = ws.max_column
    asin_cols = []
    for c in range(4, max_col + 1):
        asin = ws.cell(row=header_row, column=c).value
        if asin:
            asin_cols.append((c, str(asin).strip()))

    records = []
    r = header_row + 1
    blank_streak = 0
    while r <= ws.max_row and blank_streak < 3:
        kw = ws.cell(row=r, column=1).value
        if kw is None or str(kw).strip() == "":
            blank_streak += 1
            r += 1
            continue
        blank_streak = 0
        sv = _clean_num(ws.cell(row=r, column=2).value)
        rel = _clean_num(ws.cell(row=r, column=3).value)
        for col_idx, asin in asin_cols:
            rank = _clean_num(ws.cell(row=r, column=col_idx).value)
            records.append({
                "keyword_phrase": str(kw).strip(),
                "search_volume": sv,
                "relevancy": rel,
                "asin": asin,
                "organic_rank": rank,
            })
        r += 1

    grid = pd.DataFrame(records)
    return {"main_keyword": main_keyword, "roots": roots, "grid": grid}


def compute_competitor_scores(grid: pd.DataFrame, rank_thresh=30, rel_thresh=0, sv_thresh=0):
    """
    For each ASIN, compute the same metrics the manual template uses:
    SV captured (rank < rank_thresh & relevancy > rel_thresh), as % of total
    relevant search volume; matching keyword count as % of total relevant
    keyword count. Then classify green/yellow/orange/red per the legend:
    VERY GOOD (strong competitor) >80% => red, GOOD >60% => orange,
    OKAY >40% => yellow, BAD <40% => green (weak competitor = opportunity).
    """
    universe = grid.drop_duplicates(subset=["keyword_phrase"])
    has_relevancy = universe["relevancy"].notna().any()
    if has_relevancy:
        universe = universe[(universe["relevancy"].fillna(0) > rel_thresh) & (universe["search_volume"].fillna(0) > sv_thresh)]
    else:
        universe = universe[universe["search_volume"].fillna(0) > sv_thresh]
    relevant_keywords = set(universe["keyword_phrase"])
    total_sv = universe["search_volume"].sum()
    total_kw = len(universe)

    rows = []
    for asin, sub in grid[grid["keyword_phrase"].isin(relevant_keywords)].groupby("asin"):
        matched = sub[(sub["organic_rank"].notna()) & (sub["organic_rank"] < rank_thresh)]
        sv_captured = matched["search_volume"].sum()
        kw_captured = len(matched)
        pct_sv = (sv_captured / total_sv * 100) if total_sv else 0
        pct_kw = (kw_captured / total_kw * 100) if total_kw else 0

        if pct_sv > 80:
            zone, color = "VERY STRONG", "🔴"
        elif pct_sv > 60:
            zone, color = "STRONG", "🟠"
        elif pct_sv > 40:
            zone, color = "MODERATE", "🟡"
        else:
            zone, color = "WEAK", "🟢"

        rows.append({
            "asin": asin,
            "sv_captured": sv_captured,
            "pct_sv": round(pct_sv, 1),
            "kw_captured": kw_captured,
            "pct_kw": round(pct_kw, 1),
            "zone": zone,
            "color": color,
        })

    return pd.DataFrame(rows).sort_values("pct_sv", ascending=False), total_sv, total_kw


def niche_verdict(scores: pd.DataFrame):
    red = (scores["color"] == "🔴").sum()
    orange = (scores["color"] == "🟠").sum()
    yellow = (scores["color"] == "🟡").sum()
    green = (scores["color"] == "🟢").sum()

    if red > 4:
        return "🔴 NO GO — 4 se zyada bohot strong (red) competitors hain.", red, orange, yellow, green
    if orange > 5:
        return "🟠 HIGH RISK — 5 se zyada strong (orange) competitors hain.", red, orange, yellow, green
    if green >= 2 and orange <= 2 and red == 0:
        if yellow <= 1:
            return "🟢 EXCELLENT — zyadatar competitors weak hain, kam competition.", red, orange, yellow, green
        return "🟢 GOOD — weak/moderate competitors ka mix hai, 1-2 strong competitor.", red, orange, yellow, green
    return "🟡 MIXED — dhyan se dekhein, profit margin aur product variations bhi check karein.", red, orange, yellow, green

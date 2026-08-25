"""
export_report.py
------------------
Builds styled Excel reports.

1. build_cerebro_report      -> keyword table + competitor/brand comparison
                                 + blank rank-grid template (Keyword Research page)
2. build_master_kw_report    -> colored "Master Keyword Research" report:
                                 competitor scorecard + full keyword x ASIN
                                 organic rank grid, cells color-coded by rank
                                 (Master Keyword Research page)
"""

from io import BytesIO

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

HEADER_FILL = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
HEADER_FONT = Font(color="FFFFFF", bold=True)
TITLE_FONT = Font(bold=True, size=14)
SUBTLE_FILL = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
THIN = Side(style="thin", color="B7B7B7")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

# rank-grid cell colors
RANK_GOOD_FILL = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")   # green - rank < 10
RANK_OK_FILL = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")     # yellow - rank < 30
RANK_BAD_FILL = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")    # red - rank >= 30
RANK_GOOD_FONT = Font(color="006100")
RANK_OK_FONT = Font(color="9C6500")
RANK_BAD_FONT = Font(color="9C0006")

# scorecard zone colors (matches the 🔴🟠🟡🟢 emoji already in the "color" column)
ZONE_FILL = {
    "🔴": PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid"),
    "🟠": PatternFill(start_color="FFE0B2", end_color="FFE0B2", fill_type="solid"),
    "🟡": PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid"),
    "🟢": PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid"),
}


def _style_header_row(ws, row, col_start, col_end):
    for c in range(col_start, col_end + 1):
        cell = ws.cell(row=row, column=c)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = BORDER


def _autofit(ws, col_start, col_end, min_width=10, max_width=42):
    for c in range(col_start, col_end + 1):
        letter = get_column_letter(c)
        longest = min_width
        for cell in ws[letter]:
            if cell.value is not None:
                longest = max(longest, len(str(cell.value)) + 2)
        ws.column_dimensions[letter].width = min(longest, max_width)


# ---------------------------------------------------------------------------
# Report 1: Keyword Research page (Cerebro keyword list + brand comparison)
# ---------------------------------------------------------------------------
def build_cerebro_report(niche, keyword_df, market_df=None):
    """
    keyword_df: DataFrame with columns keyword_phrase, match_type, relevancy,
                cerebro_iq_score, search_volume (from cerebro_keywords table)
    market_df:  DataFrame with columns asin, brand, price, sales, revenue
                (from market_analysis table) — optional, used for the
                competitor comparison block and rank-grid column headers.
    Returns: BytesIO of the .xlsx file.
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Keyword Research"

    total_kw = len(keyword_df)
    total_sv = int(keyword_df["search_volume"].fillna(0).sum())

    # --- Title block ---
    ws["A1"] = niche
    ws["A1"].font = TITLE_FONT
    ws["A2"] = "Total Keywords"
    ws["B2"] = total_kw
    ws["A3"] = "Total Search Volume"
    ws["B3"] = total_sv
    for cell in ("A2", "A3"):
        ws[cell].font = Font(bold=True)

    # --- Section 1: Keyword table ---
    header_row = 5
    headers = ["Keyword Phrase", "Match Type", "Relevancy", "Cerebro IQ Score", "Search Volume"]
    for i, h in enumerate(headers, start=1):
        ws.cell(row=header_row, column=i, value=h)
    _style_header_row(ws, header_row, 1, len(headers))

    kw_sorted = keyword_df.sort_values("search_volume", ascending=False, na_position="last")
    r = header_row + 1
    for _, row in kw_sorted.iterrows():
        ws.cell(row=r, column=1, value=row.get("keyword_phrase"))
        ws.cell(row=r, column=2, value=row.get("match_type"))
        ws.cell(row=r, column=3, value=row.get("relevancy"))
        ws.cell(row=r, column=4, value=row.get("cerebro_iq_score"))
        ws.cell(row=r, column=5, value=row.get("search_volume"))
        for c in range(1, 6):
            ws.cell(row=r, column=c).border = BORDER
        if r % 2 == 0:
            for c in range(1, 6):
                ws.cell(row=r, column=c).fill = SUBTLE_FILL
        r += 1
    last_kw_row = r - 1
    _autofit(ws, 1, 5)
    ws.freeze_panes = ws.cell(row=header_row + 1, column=1).coordinate

    # --- Section 2: Competitor / brand comparison block (to the right) ---
    if market_df is not None and not market_df.empty:
        block_col_start = 8  # column H
        brands = market_df.sort_values("sales", ascending=False)

        ws.cell(row=header_row, column=block_col_start, value="Metric")
        for j, (_, brow) in enumerate(brands.iterrows()):
            col = block_col_start + 1 + j
            ws.cell(row=header_row, column=col, value=brow.get("asin"))
        _style_header_row(ws, header_row, block_col_start, block_col_start + len(brands))

        metrics = [
            ("Brand Name", "brand"),
            ("Price ($)", "price"),
            ("Unit Sales (Monthly)", "sales"),
            ("Revenue ($)", "revenue"),
        ]
        for m_i, (label, field) in enumerate(metrics):
            row_i = header_row + 1 + m_i
            ws.cell(row=row_i, column=block_col_start, value=label).font = Font(bold=True)
            for j, (_, brow) in enumerate(brands.iterrows()):
                col = block_col_start + 1 + j
                ws.cell(row=row_i, column=col, value=brow.get(field))
            for c in range(block_col_start, block_col_start + len(brands) + 1):
                ws.cell(row=row_i, column=c).border = BORDER

        _autofit(ws, block_col_start, block_col_start + len(brands))

        # --- Section 3: Keyword x ASIN organic rank grid (template — manual fill) ---
        grid_start_row = header_row + len(metrics) + 3
        ws.cell(row=grid_start_row - 1, column=block_col_start,
                value="Organic Rank by ASIN — fill in manually from your rank checks").font = Font(bold=True, italic=True)

        ws.cell(row=grid_start_row, column=1, value="Keyword Phrase")
        for j, (_, brow) in enumerate(brands.iterrows()):
            col = 2 + j
            ws.cell(row=grid_start_row, column=col, value=brow.get("asin"))
        _style_header_row(ws, grid_start_row, 1, 1 + len(brands))

        gr = grid_start_row + 1
        for _, krow in kw_sorted.iterrows():
            ws.cell(row=gr, column=1, value=krow.get("keyword_phrase"))
            ws.cell(row=gr, column=1).border = BORDER
            for j in range(len(brands)):
                cell = ws.cell(row=gr, column=2 + j)
                cell.border = BORDER
                cell.fill = PatternFill(start_color="FFF9E6", end_color="FFF9E6", fill_type="solid")  # blank-to-fill highlight
            gr += 1

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


# ---------------------------------------------------------------------------
# Report 2: Master Keyword Research page (competitor scorecard + rank grid)
# ---------------------------------------------------------------------------
def build_master_kw_report(niche, main_keyword, grid, scores, total_sv, total_kw, verdict):
    """
    niche:         niche name (str)
    main_keyword:  main/root keyword (str, may be None)
    grid:          long-format DataFrame: keyword_phrase, search_volume,
                   relevancy, asin, organic_rank
    scores:        DataFrame from compute_competitor_scores: asin,
                   sv_captured, pct_sv, kw_captured, pct_kw, zone, color
    total_sv:      total relevant search volume (float)
    total_kw:      total relevant keyword count (int)
    verdict:       verdict string (e.g. "🟢 GOOD — ...")
    Returns: BytesIO of the .xlsx file.
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Master Keyword Research"

    # --- Title block ---
    ws["A1"] = niche
    ws["A1"].font = TITLE_FONT
    ws["A2"] = "Main Keyword"
    ws["B2"] = main_keyword or ""
    ws["A3"] = "Total Relevant Keywords"
    ws["B3"] = int(total_kw)
    ws["A4"] = "Total Relevant Search Volume"
    ws["B4"] = int(total_sv) if pd.notna(total_sv) else 0
    ws["A5"] = "Verdict"
    ws["B5"] = verdict
    for cell in ("A2", "A3", "A4", "A5"):
        ws[cell].font = Font(bold=True)
    ws["B5"].font = Font(bold=True)

    # --- Section 1: Competitor scorecard ---
    header_row = 7
    headers = ["ASIN", "SV Captured", "% of Total SV", "Keywords Captured", "% of Total KW", "Zone", ""]
    for i, h in enumerate(headers, start=1):
        ws.cell(row=header_row, column=i, value=h)
    _style_header_row(ws, header_row, 1, len(headers))

    r = header_row + 1
    for _, row in scores.iterrows():
        ws.cell(row=r, column=1, value=row.get("asin"))
        ws.cell(row=r, column=2, value=row.get("sv_captured"))
        ws.cell(row=r, column=3, value=row.get("pct_sv"))
        ws.cell(row=r, column=4, value=row.get("kw_captured"))
        ws.cell(row=r, column=5, value=row.get("pct_kw"))
        ws.cell(row=r, column=6, value=row.get("zone"))
        ws.cell(row=r, column=7, value=row.get("color"))
        fill = ZONE_FILL.get(row.get("color"))
        for c in range(1, 8):
            cell = ws.cell(row=r, column=c)
            cell.border = BORDER
            if fill:
                cell.fill = fill
        r += 1
    _autofit(ws, 1, 7)

    # --- Section 2: Keyword x ASIN organic rank grid ---
    grid_start_row = r + 2
    ws.cell(row=grid_start_row - 1, column=1,
            value="Keyword x ASIN Organic Rank Grid").font = Font(bold=True, italic=True)

    asins = sorted(grid["asin"].dropna().unique())
    pivot = grid.pivot_table(index="keyword_phrase", columns="asin", values="organic_rank", aggfunc="first")
    meta = grid.drop_duplicates(subset=["keyword_phrase"]).set_index("keyword_phrase")[["search_volume", "relevancy"]]
    display_df = meta.join(pivot).sort_values("search_volume", ascending=False)

    grid_headers = ["Keyword Phrase", "Search Volume", "Relevancy"] + list(asins)
    for i, h in enumerate(grid_headers, start=1):
        ws.cell(row=grid_start_row, column=i, value=h)
    _style_header_row(ws, grid_start_row, 1, len(grid_headers))

    gr = grid_start_row + 1
    for kw, row in display_df.iterrows():
        ws.cell(row=gr, column=1, value=kw)
        ws.cell(row=gr, column=2, value=row.get("search_volume"))
        ws.cell(row=gr, column=3, value=row.get("relevancy"))
        for c in range(1, 4):
            ws.cell(row=gr, column=c).border = BORDER

        for j, asin in enumerate(asins):
            col = 4 + j
            rank = row.get(asin)
            cell = ws.cell(row=gr, column=col)
            cell.border = BORDER
            if pd.isna(rank):
                cell.value = "-"
            else:
                cell.value = int(rank)
                if rank < 10:
                    cell.fill = RANK_GOOD_FILL
                    cell.font = RANK_GOOD_FONT
                elif rank < 30:
                    cell.fill = RANK_OK_FILL
                    cell.font = RANK_OK_FONT
                else:
                    cell.fill = RANK_BAD_FILL
                    cell.font = RANK_BAD_FONT
        gr += 1

    _autofit(ws, 1, 3 + len(asins))
    ws.freeze_panes = ws.cell(row=grid_start_row + 1, column=4).coordinate

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf
